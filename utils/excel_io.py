"""
Excel import/export utilities using openpyxl.
"""
import os
from openpyxl import load_workbook, Workbook
from app.models import Sheet, SheetColumn, SheetRow
from app.extensions import db
from flask import current_app


def import_excel(file_obj, stored_path, excel_file_record):
    """
    Read an uploaded Excel file, parse all sheets into Sheet/SheetColumn/SheetRow.

    Args:
        file_obj: file-like object from the uploaded form
        stored_path: where the file is saved on disk
        excel_file_record: ExcelFile ORM instance (already committed)

    Returns:
        list of Sheet instances created
    """
    file_obj.save(stored_path)
    wb = load_workbook(stored_path, data_only=True)

    created_sheets = []

    for sheet_order, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]

        # Determine row bounds
        rows = list(ws.iter_rows())
        if not rows:
            continue

        # First row = headers
        header_row = rows[0]
        headers = []
        for col_idx, cell in enumerate(header_row):
            if cell.value is not None:
                headers.append((col_idx, str(cell.value).strip()))

        if not headers:
            continue

        sheet = Sheet(
            file_id=excel_file_record.id,
            sheet_name=sheet_name,
            sheet_order=sheet_order,
        )
        db.session.add(sheet)
        db.session.flush()  # get sheet.id

        # Create SheetColumn records
        for col_order, (col_idx, label) in enumerate(headers):
            col_key = label  # can be normalized later if needed
            col = SheetColumn(
                sheet_id=sheet.id,
                column_key=col_key,
                column_label=label,
                column_order=col_order,
            )
            db.session.add(col)

        # Create SheetRow records (row 1 onward = data)
        for row_order, row in enumerate(rows[1:]):
            row_data = {}
            for col_order, (col_idx, label) in enumerate(headers):
                cell = row[col_idx] if col_idx < len(row) else None
                row_data[label] = str(cell.value) if cell is not None and cell.value is not None else ''

            sr = SheetRow(
                sheet_id=sheet.id,
                row_order=row_order,
                data=row_data,
            )
            db.session.add(sr)

        created_sheets.append(sheet)

    wb.close()
    db.session.commit()
    return created_sheets


def export_excel(excel_file_record):
    """
    Rebuild an Excel file from database records.

    Returns:
        path to the generated .xlsx file (string)
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets = Sheet.query.filter_by(file_id=excel_file_record.id)\
        .order_by(Sheet.sheet_order).all()

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet.sheet_name[:31])  # sheet name max 31 chars

        # Get columns ordered
        columns = SheetColumn.query.filter_by(sheet_id=sheet.id)\
            .order_by(SheetColumn.column_order).all()

        # Write header row
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col.column_label)

        # Write data rows
        rows = SheetRow.query.filter_by(sheet_id=sheet.id)\
            .order_by(SheetRow.row_order).all()

        for row in rows:
            excel_row = row.row_order + 2  # +2: 1-indexed + header row
            for col_idx, col in enumerate(columns, start=1):
                value = row.data.get(col.column_key, '') if isinstance(row.data, dict) else ''
                ws.cell(row=excel_row, column=col_idx, value=value)

    export_dir = current_app.config.get('UPLOAD_FOLDER', 'uploads')
    export_path = os.path.join(export_dir, f'export_{excel_file_record.id}.xlsx')
    wb.save(export_path)
    wb.close()
    return export_path
